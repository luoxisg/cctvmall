from __future__ import annotations

import argparse
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PRIORITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
ACTIVE_STATUSES = {"Open", "In Progress", "Blocked", "Waiting External", "Review"}
STATUS_FILL = {
    "Open": "FFF8CC",
    "In Progress": "DDEBF7",
    "Blocked": "F8CBAD",
    "Done": "E2F0D9",
}


@dataclass
class ActionRecord:
    file_name: str
    note_link: str
    action_title: str
    site_name: str
    status: str
    completed: bool
    priority: str
    due_date: date | None
    owner_org: str
    owner_person: str
    related_submission: str
    source_ref: str
    blocker: str
    depends_on: str

    @property
    def effective_status(self) -> str:
        if self.completed or self.status == "Done":
            return "Done"
        return self.status or "Open"

    @property
    def is_active(self) -> bool:
        return self.effective_status in ACTIVE_STATUSES


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export current action plan to Excel.")
    parser.add_argument(
        "--as-of",
        help="As-of date in YYYY-MM-DD format. Defaults to today.",
    )
    return parser.parse_args()


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def extract_frontmatter(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    if not text.startswith("---\n"):
        return {}
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}
    return yaml.safe_load(parts[1]) or {}


def load_actions(actions_dir: Path) -> list[ActionRecord]:
    records: list[ActionRecord] = []
    for path in sorted(actions_dir.glob("*.md")):
        data = extract_frontmatter(path)
        record = ActionRecord(
            file_name=path.name,
            note_link=str(path.relative_to(actions_dir.parent.parent)).replace("\\", "/"),
            action_title=str(data.get("action_title", "")).strip(),
            site_name=str(data.get("site_name", "")).strip(),
            status=str(data.get("status", "Open")).strip() or "Open",
            completed=bool(data.get("completed", False)),
            priority=str(data.get("priority", "Medium")).strip() or "Medium",
            due_date=parse_date(data.get("due_date")),
            owner_org=str(data.get("owner_org", "")).strip(),
            owner_person=str(data.get("owner_person", "")).strip(),
            related_submission=str(data.get("related_submission", "")).strip(),
            source_ref=str(data.get("source_ref", "")).strip(),
            blocker=str(data.get("blocker", "")).strip(),
            depends_on=str(data.get("depends_on", "")).strip(),
        )
        records.append(record)
    return records


def sort_key(record: ActionRecord) -> tuple[Any, ...]:
    return (
        record.due_date or date.max,
        PRIORITY_ORDER.get(record.priority, 99),
        record.site_name.lower(),
        record.action_title.lower(),
    )


def style_sheet(ws, status_column_index: int | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2):
        if status_column_index is not None and len(row) >= status_column_index:
            status_value = str(row[status_column_index - 1].value or "")
            fill_color = STATUS_FILL.get(status_value)
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for idx, column_cells in enumerate(ws.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 48)


def add_sheet(ws, records: list[ActionRecord]) -> None:
    ws.append(
        [
            "Action Title",
            "Site",
            "Status",
            "Priority",
            "Due Date",
            "Owner Org",
            "Owner Person",
            "Related Submission",
            "Blocker",
            "Depends On",
            "Source Ref",
            "Completed",
            "Action Note",
        ]
    )
    for record in records:
        ws.append(
            [
                record.action_title,
                record.site_name,
                record.effective_status,
                record.priority,
                record.due_date.isoformat() if record.due_date else "",
                record.owner_org,
                record.owner_person,
                record.related_submission,
                record.blocker,
                record.depends_on,
                record.source_ref,
                "Yes" if record.completed else "No",
                record.note_link,
            ]
        )
    style_sheet(ws, status_column_index=3)


def add_summary_sheet(ws, records: list[ActionRecord], as_of: date) -> None:
    active = [record for record in records if record.is_active]
    done = [record for record in records if not record.is_active]
    status_counts = {
        "Open": sum(1 for record in active if record.effective_status == "Open"),
        "In Progress": sum(1 for record in active if record.effective_status == "In Progress"),
        "Blocked": sum(1 for record in active if record.effective_status == "Blocked"),
        "Done": len(done),
    }

    ws.append(["Metric", "Value"])
    ws.append(["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["As Of Date", as_of.isoformat()])
    ws.append(["Xjera Documents Due to Univers", "2026-05-04"])
    ws.append(["HDB Submission Target by Univers", "2026-05-06"])
    ws.append(["Total Active Actions", len(active)])
    ws.append(["Open Actions", status_counts["Open"]])
    ws.append(["In Progress Actions", status_counts["In Progress"]])
    ws.append(["Blocked Actions", status_counts["Blocked"]])
    ws.append(["Completed Actions", status_counts["Done"]])
    ws.append(["Priority Sites", "Loyang Point; Canberra Plaza"])
    style_sheet(ws)


def main() -> None:
    args = parse_args()
    as_of = date.fromisoformat(args.as_of) if args.as_of else date.today()

    script_path = Path(__file__).resolve()
    vault_root = script_path.parent.parent
    actions_dir = vault_root / "08_Bases" / "Actions"
    export_dir = vault_root / "07_Attachments" / "Excel"
    export_dir.mkdir(parents=True, exist_ok=True)

    records = load_actions(actions_dir)
    active_records = sorted([record for record in records if record.is_active], key=sort_key)
    done_records = sorted([record for record in records if not record.is_active], key=sort_key)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    add_summary_sheet(summary_ws, records, as_of)

    action_ws = workbook.create_sheet("Action Plan")
    add_sheet(action_ws, active_records)

    done_ws = workbook.create_sheet("Completed")
    add_sheet(done_ws, done_records)

    dated_name = f"Action Plan - {as_of.isoformat()}.xlsx"
    dated_path = export_dir / dated_name
    current_path = export_dir / "Action Plan - Current.xlsx"

    workbook.save(dated_path)
    shutil.copyfile(dated_path, current_path)

    print(str(dated_path))
    print(str(current_path))


if __name__ == "__main__":
    main()
